# -*- coding: utf-8 -*-
"""export_excel.py — 将群用户性别推理结果导出为 Excel（是/否，无符号）

用法：
  python train/export_excel.py --csv outputs/群1用户性别推理.csv --group 826904606 \
      --out outputs/群1用户性别推理.xlsx
"""
import argparse
import csv
import json
import os
import re
import sqlite3

from openpyxl import Workbook

# Excel 不允许的字符：控制符（除 \t \n \r）及非字符码点
_ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\ufffe\uffff]')


def clean(s):
    return _ILLEGAL.sub('', s) if isinstance(s, str) else s
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CONF_CN = {'high': '高', 'borderline': '中（临界）', 'low-data': '低（样本不足）'}
PRED_CN = {'male': '男', 'female': '女'}
SEX_CN = {'male': '男', 'female': '女'}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--csv', required=True)
    ap.add_argument('--group', type=int, required=True)
    ap.add_argument('--out', required=True)
    ap.add_argument('--min-msgs', type=int, default=5)
    args = ap.parse_args()

    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    with open(os.path.join(root, 'config', 'config.json'), encoding='utf-8-sig') as f:
        db_path = json.load(f).get('database', 'data/qqchat.db')
    if not os.path.isabs(db_path):
        db_path = os.path.join(root, db_path)

    # 已打分用户（来自 score_group.py 的 CSV）
    scored = {}
    with open(args.csv, encoding='utf-8') as f:
        for r in csv.DictReader(f):
            scored[int(r['user_id'])] = r

    # 群内全部用户：发言数、昵称、群名片
    conn = sqlite3.connect(f'file:{db_path}?mode=ro', uri=True)
    conn.row_factory = sqlite3.Row
    users = {}
    for r in conn.execute("""
        SELECT user_id, COUNT(*) c FROM messages
        WHERE scene='group' AND peer_id=? GROUP BY user_id""", (args.group,)):
        users[r['user_id']] = {'n': r['c']}
    for uid in users:
        nick = conn.execute("SELECT nickname FROM user_profiles WHERE user_id=?", (uid,)).fetchone()
        users[uid]['nickname'] = nick['nickname'] if nick and nick['nickname'] else ''
        card = conn.execute("""
            SELECT card FROM messages WHERE peer_id=? AND user_id=? AND card IS NOT NULL AND card != ''
            ORDER BY time DESC LIMIT 1""", (args.group, uid)).fetchone()
        users[uid]['card'] = card['card'] if card else ''
    conn.close()

    train_set = set()
    val_set = set()
    for fname, s in [('data/train.jsonl', train_set), ('data/val.jsonl', val_set)]:
        p = os.path.join(root, fname)
        if os.path.exists(p):
            with open(p, encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line:
                        s.add(json.loads(line)['user_id'])

    main_rows = []
    skip_rows = []
    for uid, info in users.items():
        base = {'user_id': uid, 'nickname': info['nickname'], 'card': info['card'], 'n': info['n']}
        if uid in scored:
            r = scored[uid]
            base.update({
                'mean': r['mean'], 'pred': PRED_CN.get(r['pred'], r['pred']),
                'conf': CONF_CN.get(r['conf'], r['conf']),
                'in_train': '是' if uid in train_set else '否',
                'in_val': '是' if uid in val_set else '否',
                'label': SEX_CN.get(r['label'], '') if r.get('label') else '',
            })
            main_rows.append(base)
        else:
            base.update({'mean': '', 'pred': '样本不足', 'conf': '', 'in_train': '否', 'in_val': '否', 'label': ''})
            skip_rows.append(base)

    main_rows.sort(key=lambda r: -int(r['n']))
    skip_rows.sort(key=lambda r: -int(r['n']))

    headers = ['QQ号', '全局昵称', '群昵称', '发言数', '女概率均值', '结论', '置信度',
               '是否训练集', '是否测试集', '人工标签']
    widths = [14, 22, 30, 9, 12, 10, 14, 11, 11, 10]

    wb = Workbook()
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill_sheet(ws, rows):
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
        for row in rows:
            ws.append([clean(str(row['user_id'])), clean(row['nickname']), clean(row['card']), int(row['n']),
                       float(row['mean']) if row['mean'] != '' else '',
                       row['pred'], row['conf'], row['in_train'], row['in_val'], row['label']])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border
                if cell.column in (1, 4, 5, 6, 7, 8, 9, 10):
                    cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'

    ws1 = wb.active
    ws1.title = '推理结果'
    fill_sheet(ws1, main_rows)

    ws2 = wb.create_sheet('样本不足')
    fill_sheet(ws2, skip_rows)

    os.makedirs(os.path.dirname(args.out) or '.', exist_ok=True)
    wb.save(args.out)
    print(f'[完成] 推理结果 {len(main_rows)} 人 + 样本不足 {len(skip_rows)} 人 → {args.out}')


if __name__ == '__main__':
    main()
